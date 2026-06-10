#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 生成工具 - 支持中文、格式化
支持 openpyxl 和 xlsxwriter 两种引擎
"""

import os
from typing import List, Dict, Any, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlsxwriter
    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False

try:
    from .fonts import font_manager, TEST_STRINGS
except ImportError:
    from fonts import font_manager, TEST_STRINGS


class ExcelGenerator:
    """Excel 生成器"""
    
    def __init__(self, engine='openpyxl'):
        self.engine = engine
        self.font_name = font_manager.get_font_name()
        
        if engine == 'openpyxl' and not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl 未安装，请运行: pip install openpyxl")
        if engine == 'xlsxwriter' and not XLSXWRITER_AVAILABLE:
            raise ImportError("xlsxwriter 未安装，请运行: pip install xlsxwriter")
    
    def create_from_data(self, output_path: str, data: List[List[Any]], 
                         headers: Optional[List[str]] = None,
                         sheet_name: str = "Sheet1",
                         title: Optional[str] = None) -> str:
        """
        从数据创建Excel
        
        参数:
            output_path: 输出路径
            data: 二维数据列表
            headers: 列标题（可选）
            sheet_name: 工作表名称
            title: 表格标题（可选）
        """
        if self.engine == 'openpyxl':
            return self._create_with_openpyxl(output_path, data, headers, sheet_name, title)
        else:
            return self._create_with_xlsxwriter(output_path, data, headers, sheet_name, title)
    
    def _create_with_openpyxl(self, output_path: str, data: List[List[Any]],
                              headers: Optional[List[str]],
                              sheet_name: str, title: Optional[str]) -> str:
        """使用openpyxl创建"""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 字体设置
        default_font = Font(name=self.font_name, size=11)
        header_font = Font(name=self.font_name, size=12, bold=True, color="FFFFFF")
        title_font = Font(name=self.font_name, size=14, bold=True)
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 1
        
        # 标题
        if title:
            num_cols = len(headers) if headers else len(data[0]) if data else 1
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row = 2
        
        # 表头
        if headers:
            for col_idx, header in enumerate(headers):
                cell = ws.cell(row=current_row, column=col_idx + 1, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            current_row += 1
        
        # 数据
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(row=current_row + row_idx, column=col_idx + 1, value=value)
                cell.font = default_font
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = thin_border
        
        # 自动调整列宽
        num_cols = len(headers) if headers else len(data[0]) if data else 1
        for col_idx in range(1, num_cols + 1):
            column = get_column_letter(col_idx)
            max_length = 0
            for row in ws[column]:
                try:
                    if row.value:
                        cell_length = len(str(row.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        wb.save(output_path)
        return output_path
    
    def _create_with_xlsxwriter(self, output_path: str, data: List[List[Any]],
                                headers: Optional[List[str]],
                                sheet_name: str, title: Optional[str]) -> str:
        """使用xlsxwriter创建"""
        workbook = xlsxwriter.Workbook(output_path)
        worksheet = workbook.add_worksheet(sheet_name)
        
        # 格式定义
        title_format = workbook.add_format({
            'font_name': self.font_name,
            'font_size': 14,
            'bold': True,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        header_format = workbook.add_format({
            'font_name': self.font_name,
            'font_size': 12,
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'bg_color': '#4472C4',
            'font_color': 'white',
            'border': 1
        })
        
        data_format = workbook.add_format({
            'font_name': self.font_name,
            'font_size': 11,
            'align': 'left',
            'valign': 'vcenter',
            'border': 1
        })
        
        row = 0
        num_cols = len(headers) if headers else len(data[0]) if data else 1
        
        # 标题
        if title:
            worksheet.merge_range(0, 0, 0, num_cols - 1, title, title_format)
            worksheet.set_row(0, 30)
            row = 1
        
        # 表头
        if headers:
            for col_idx, header in enumerate(headers):
                worksheet.write(row, col_idx, header, header_format)
            row += 1
        
        # 数据
        for row_data in data:
            for col_idx, value in enumerate(row_data):
                worksheet.write(row, col_idx, value, data_format)
            row += 1
        
        # 设置列宽
        for i in range(num_cols):
            worksheet.set_column(i, i, 15)
        
        workbook.close()
        return output_path


def create_simple_excel(output_path: str, data: List[List[Any]], 
                        headers: Optional[List[str]] = None,
                        sheet_name: str = "Sheet1",
                        title: Optional[str] = None,
                        engine: str = 'openpyxl') -> str:
    """
    快速创建Excel
    
    参数:
        output_path: 输出路径
        data: 数据列表
        headers: 列标题
        sheet_name: 工作表名称
        title: 表格标题
        engine: 'openpyxl' 或 'xlsxwriter'
    
    例子:
        create_simple_excel('test.xlsx', 
            [['张三', 85], ['李四', 92]],
            headers=['姓名', '分数'],
            title='成绩单'
        )
    """
    generator = ExcelGenerator(engine=engine)
    return generator.create_from_data(output_path, data, headers, sheet_name, title)


def create_test_excel(output_path: str, engine: str = 'openpyxl') -> str:
    """创建测试Excel，验证中文显示"""
    generator = ExcelGenerator(engine=engine)
    
    data = [
        [TEST_STRINGS['chinese'], '通过'],
        [TEST_STRINGS['punctuation'], '通过'],
        [TEST_STRINGS['math'], '通过'],
        [TEST_STRINGS['symbols'], '通过'],
        [TEST_STRINGS['arrows'], '通过'],
        [TEST_STRINGS['currency'], '通过'],
        [TEST_STRINGS['mixed'], '通过'],
    ]
    
    headers = ['测试内容', '状态']
    
    return generator.create_from_data(output_path, data, headers, '测试数据', '中文Excel测试')


if __name__ == "__main__":
    print("测试Excel生成...")
    
    try:
        path = create_test_excel('/tmp/test_chinese_openpyxl.xlsx', engine='openpyxl')
        print(f"✓ openpyxl测试Excel已生成: {path}")
    except Exception as e:
        print(f"✗ openpyxl测试失败: {e}")
    
    try:
        path = create_test_excel('/tmp/test_chinese_xlsxwriter.xlsx', engine='xlsxwriter')
        print(f"✓ xlsxwriter测试Excel已生成: {path}")
    except Exception as e:
        print(f"✗ xlsxwriter测试失败: {e}")
