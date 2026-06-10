#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
中文Excel生成工具 - 已配置好中文字体
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import xlsxwriter

# ============== 字体配置 ==============
# 常用中文字体名称（系统字体或WPS/Office字体）
CHINESE_FONTS = {
    'noto_sans': 'Noto Sans CJK SC',
    'microsoft_yahei': 'Microsoft YaHei',
    'simhei': 'SimHei',
    'simsun': 'SimSun',
    'arial_unicode': 'Arial Unicode MS',
}

DEFAULT_FONT = CHINESE_FONTS['noto_sans']


# ============== openpyxl 方式 ==============

def create_excel_openpyxl(output_path, data, sheet_name="Sheet1", title=None):
    """
    使用openpyxl创建Excel - 适合复杂格式
    
    参数:
        output_path: 输出文件路径
        data: 数据列表，每行是一个列表
        sheet_name: 工作表名称
        title: 标题（可选）
    
    例子:
        data = [
            ["姓名", "年龄", "分数"],
            ["张三", 20, 85],
            ["李四", 21, 92],
        ]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 设置默认字体
    default_font = Font(name=DEFAULT_FONT, size=11)
    
    current_row = 1
    
    # 添加标题
    if title:
        ws.merge_cells(f'A1:{get_column_letter(len(data[0]))}1')
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(name=DEFAULT_FONT, size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row = 2
    
    # 填充数据
    for row_idx, row_data in enumerate(data):
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=current_row + row_idx, column=col_idx + 1, value=value)
            cell.font = default_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 自动调整列宽
    for col_idx in range(1, len(data[0]) + 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for row in ws[column]:
            try:
                if row.value:
                    cell_length = len(str(row.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(output_path)
    print(f"Excel已生成: {output_path}")
    return output_path


def create_styled_excel(output_path, data, headers, sheet_name="数据表"):
    """
    创建带样式的Excel - 标题行突出显示
    
    参数:
        output_path: 输出路径
        data: 数据列表（不含头）
        headers: 标题行列表
        sheet_name: 工作表名称
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 样式定义
    header_font = Font(name=DEFAULT_FONT, size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    data_font = Font(name=DEFAULT_FONT, size=11)
    data_alignment = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入标题行
    for col_idx, header in enumerate(headers):
        cell = ws.cell(row=1, column=col_idx + 1, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入数据
    for row_idx, row_data in enumerate(data):
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=row_idx + 2, column=col_idx + 1, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # 自动调整列宽
    for col_idx in range(1, len(headers) + 1):
        column = get_column_letter(col_idx)
        ws.column_dimensions[column].auto_size = True
    
    wb.save(output_path)
    print(f"Excel已生成: {output_path}")
    return output_path


# ============== xlsxwriter 方式 (推荐用于大文件) ==============

def create_excel_xlsxwriter(output_path, data, sheet_name="Sheet1", title=None):
    """
    使用xlsxwriter创建Excel - 适合大文件和复杂格式
    
    参数同create_excel_openpyxl
    """
    workbook = xlsxwriter.Workbook(output_path)
    worksheet = workbook.add_worksheet(sheet_name)
    
    # 格式定义
    title_format = workbook.add_format({
        'font_name': DEFAULT_FONT,
        'font_size': 14,
        'bold': True,
        'align': 'center',
        'valign': 'vcenter'
    })
    
    header_format = workbook.add_format({
        'font_name': DEFAULT_FONT,
        'font_size': 12,
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#4472C4',
        'font_color': 'white'
    })
    
    data_format = workbook.add_format({
        'font_name': DEFAULT_FONT,
        'font_size': 11,
        'align': 'left',
        'valign': 'vcenter'
    })
    
    row = 0
    
    # 添加标题
    if title:
        worksheet.merge_range(0, 0, 0, len(data[0]) - 1, title, title_format)
        worksheet.set_row(0, 30)
        row = 1
    
    # 填充数据
    for row_idx, row_data in enumerate(data):
        for col_idx, value in enumerate(row_data):
            if row_idx == 0 and not title:
                # 第一行作为标题行
                worksheet.write(row + row_idx, col_idx, value, header_format)
            else:
                worksheet.write(row + row_idx, col_idx, value, data_format)
    
    # 自动调整列宽
    for i in range(len(data[0])):
        worksheet.set_column(i, i, 15)
    
    workbook.close()
    print(f"Excel已生成: {output_path}")
    return output_path


# ============== 示例/测试 ==============
if __name__ == "__main__":
    # 测试数据
    test_data = [
        ["姓名", "年龄", "考试科目", "分数", "备注"],
        ["张三", 20, "生理学", 85, "良好"],
        ["李四", 21, "生理学", 92, "优秀"],
        ["王五", 19, "解剖学", 78, "及格"],
    ]
    
    # 测试openpyxl
    create_excel_openpyxl(
        "/tmp/test_chinese_openpyxl.xlsx",
        test_data,
        title="考研成绩测试表"
    )
    
    # 测试带样式的Excel
    create_styled_excel(
        "/tmp/test_chinese_styled.xlsx",
        test_data[1:],  # 数据
        test_data[0],   # 标题行
        sheet_name="成绩表"
    )
    
    # 测试xlsxwriter
    create_excel_xlsxwriter(
        "/tmp/test_chinese_xlsxwriter.xlsx",
        test_data,
        title="考研成绩测试表"
    )
    
    print("\n测试Excel文件已生成:")
    print("  - /tmp/test_chinese_openpyxl.xlsx")
    print("  - /tmp/test_chinese_styled.xlsx")
    print("  - /tmp/test_chinese_xlsxwriter.xlsx")
