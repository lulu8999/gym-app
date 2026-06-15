#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel生成器 - 支持中文表头和格式化
复制此文件到你的项目中使用
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 导入字体配置
from fonts import get_font_name, get_test_data

def create_excel(output_path, data, headers=None, title=None, sheet_name="Sheet1"):
    """
    创建中文Excel
    
    参数:
        output_path: 输出文件路径
        data: 数据列表，每行是一个列表
        headers: 标题行列表（可选）
        title: 文档标题（可选）
        sheet_name: 工作表名称
    
    返回:
        output_path: 输出文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    font_name = get_font_name()
    
    # 样式定义
    title_font = Font(name=font_name, size=14, bold=True)
    header_font = Font(name=font_name, size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    data_font = Font(name=font_name, size=11)
    
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 1
    
    # 添加文档标题
    if title:
        num_cols = len(headers) if headers else (len(data[0]) if data else 1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = title_font
        title_cell.alignment = center_align
        current_row = 2
    
    # 添加表头
    if headers:
        for col_idx, header in enumerate(headers):
            cell = ws.cell(row=current_row, column=col_idx + 1, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        current_row += 1
    
    # 填充数据
    for row_idx, row_data in enumerate(data):
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=current_row + row_idx, column=col_idx + 1, value=value)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border
    
    # 自动调整列宽
    num_cols = len(headers) if headers else (len(data[0]) if data else 0)
    for col_idx in range(1, num_cols + 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for row in range(1, current_row + len(data)):
            cell_value = ws.cell(row=row, column=col_idx).value
            if cell_value:
                cell_length = len(str(cell_value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(output_path)
    print(f"Excel已生成: {output_path}")
    return output_path

def create_styled_excel(output_path, data, headers, sheet_name="数据表"):
    """
    创建带样式的Excel - 标题行突出显示
    
    参数:
        output_path: 输出路径
        data: 数据列表（不含头）
        headers: 标题行列表
        sheet_name: 工作表名称
    """
    font_name = get_font_name()
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 样式定义
    header_font = Font(name=font_name, size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    data_font = Font(name=font_name, size=11)
    data_alignment = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入标题行
    for col_idx, header in enumerate(headers):
        cell = ws.cell(row=1, column=col_idx + 1, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入数据
    for row_idx, row_data in enumerate(data):
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=row_idx + 2, column=col_idx + 1, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # 自动调整列宽
    for col_idx in range(1, len(headers) + 1):
        column = get_column_letter(col_idx)
        ws.column_dimensions[column].auto_size = True
    
    wb.save(output_path)
    print(f"Excel已生成: {output_path}")
    return output_path

if __name__ == "__main__":
    # 测试数据
    test_data = [
        ["张三", 20, "生理学", 85, "良好"],
        ["李四", 21, "生理学", 92, "优秀"],
        ["王五", 19, "解剖学", 78, "及格"],
    ]
    headers = ["姓名", "年龄", "科目", "分数", "等级"]
    
    # 测试基础版
    create_excel("/tmp/test_basic.xlsx", test_data, headers=headers, title="成绩表")
    
    # 测试带样式版
    create_styled_excel("/tmp/test_styled.xlsx", test_data, headers, sheet_name="考试成绩")
    
    # 测试符号
    symbol_data = get_test_data()
    create_excel("/tmp/test_symbols.xlsx", symbol_data[1:], headers=symbol_data[0], title="符号测试")
